#!/usr/bin/env python3
try:
    import openpyxl
    
    # Load the workbook
    wb = openpyxl.load_workbook('/home/derrick/Code/testing/lab-testing/docs/work-item-export.xlsx')
    ws = wb.active
    
    print(f"Worksheet name: {ws.title}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    
    # Get headers
    print("\nHeaders:")
    headers = []
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        headers.append(header)
        print(f"Col {col}: {header}")
    
    print(f"\nTotal rows: {ws.max_row}")
    
    # Read all data
    print("\nAll work items:")
    for row in range(2, ws.max_row + 1):  # Skip header row
        row_data = {}
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row, col).value
            if headers[col-1]:  # Only if header exists
                row_data[headers[col-1]] = cell_value
        
        # Only print non-empty rows
        if any(v for v in row_data.values() if v is not None):
            print(f"\n--- Row {row} ---")
            for key, value in row_data.items():
                if value is not None:
                    print(f"{key}: {value}")
    
except ImportError:
    print("openpyxl not available, trying alternative approaches...")
    
    # Try with csv conversion using libreoffice if available
    import subprocess
    import os
    
    try:
        # Convert to CSV using LibreOffice
        subprocess.run([
            'libreoffice', '--headless', '--convert-to', 'csv', 
            '/home/derrick/Code/testing/lab-testing/docs/work-item-export.xlsx',
            '--outdir', '/tmp'
        ], check=True, capture_output=True)
        
        # Read the CSV
        csv_path = '/tmp/work-item-export.csv'
        if os.path.exists(csv_path):
            with open(csv_path, 'r') as f:
                print("CSV content:")
                print(f.read())
        else:
            print("CSV file not created")
            
    except (subprocess.CalledProcessError, FileNotFoundError):
        print("LibreOffice not available or conversion failed")
        print("Please install openpyxl: pip install openpyxl")

except Exception as e:
    print(f"Error reading Excel file: {e}")
